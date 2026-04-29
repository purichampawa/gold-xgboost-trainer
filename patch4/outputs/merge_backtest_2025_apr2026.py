"""
merge_backtest_2025_apr2026.py
==============================================
Aggregates 4 monthly backtests (set1=Jan, set2=Feb, set3=Mar, set4=Apr 2026)
into a comprehensive Excel report.

KEY DESIGN PRINCIPLES:
- Each set is an INDEPENDENT backtest (port resets to 1500 THB each month)
- Monthly results are NEVER compounded/chained — reported separately
- Cross-set aggregates are clearly labeled as "arithmetic" (not compounded)
- All metrics are computed from raw CSV data using the same logic as metrics.py
"""

from __future__ import annotations

import json
import math
import warnings
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ==========================================================
# CONFIG
# ==========================================================

BASE = Path("outputs")

SETS = {
    "Jan-2026": {
        "trades": BASE / "set1/backtests/2026-04-28_042356_trades.csv",
        "equity": BASE / "set1/backtests/2026-04-28_042356_equity.csv",
        "summary": BASE / "set1/backtests/2026-04-28_042356_summary.json",
    },
    "Feb-2026": {
        "trades": BASE / "set2/backtests/2026-04-28_042632_trades.csv",
        "equity": BASE / "set2/backtests/2026-04-28_042632_equity.csv",
        "summary": BASE / "set2/backtests/2026-04-28_042632_summary.json",
    },
    "Mar-2026": {
        "trades": BASE / "set3/backtests/2026-04-28_044052_trades.csv",
        "equity": BASE / "set3/backtests/2026-04-28_044052_equity.csv",
        "summary": BASE / "set3/backtests/2026-04-28_044052_summary.json",
    },
    "Apr-2026": {
        "trades": BASE / "set4/backtests/2026-04-28_044958_trades.csv",
        "equity": BASE / "set4/backtests/2026-04-28_044958_equity.csv",
        "summary": BASE / "set4/backtests/2026-04-28_044958_summary.json",
    },
}

STARTING_CAPITAL = 1500.0
OUTPUT_PATH = BASE / "backtest_overview_2026.xlsx"

# ==========================================================
# STYLE CONSTANTS
# ==========================================================

C_DARK_BG   = "1A1A2E"   # dark navy header
C_MID_BG    = "16213E"
C_ACCENT    = "0F3460"
C_GREEN     = "00B050"
C_RED       = "FF4B4B"
C_YELLOW    = "FFD700"
C_WHITE     = "FFFFFF"
C_LIGHT_GRAY = "F2F2F2"
C_BORDER    = "BFBFBF"

FONT_HEADER  = Font(name="Arial", bold=True, color=C_WHITE, size=11)
FONT_SUBHDR  = Font(name="Arial", bold=True, color="1A1A2E", size=10)
FONT_NORMAL  = Font(name="Arial", size=9)
FONT_BOLD    = Font(name="Arial", bold=True, size=9)
FONT_SMALL   = Font(name="Arial", size=8)
FONT_WARN    = Font(name="Arial", italic=True, color="7F7F7F", size=8)

FILL_HEADER  = PatternFill("solid", fgColor=C_DARK_BG)
FILL_SUBHDR  = PatternFill("solid", fgColor="D9E1F2")
FILL_ALT     = PatternFill("solid", fgColor="EEF3FB")
FILL_WHITE   = PatternFill("solid", fgColor=C_WHITE)
FILL_WARN    = PatternFill("solid", fgColor="FFF2CC")

THIN_SIDE  = Side(style="thin", color=C_BORDER)
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal="right", vertical="center")
ALIGN_LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ==========================================================
# HELPERS
# ==========================================================

def _safe_div(a, b, default=0.0):
    try:
        if abs(float(b)) < 1e-12:
            return default
        return float(a) / float(b)
    except Exception:
        return default


def _raw_sharpe(returns: np.ndarray) -> float:
    if len(returns) < 3:
        return 0.0
    mu, sd = returns.mean(), returns.std(ddof=1)
    return _safe_div(mu, sd)


def _annualized_sharpe(returns: np.ndarray, periods_per_year: float) -> float:
    return _raw_sharpe(returns) * math.sqrt(max(periods_per_year, 0.0))


def _deflated_sharpe(annual_sharpe: float, n_obs: int) -> float:
    if n_obs < 20:   return annual_sharpe * 0.35
    if n_obs < 60:   return annual_sharpe * 0.50
    if n_obs < 120:  return annual_sharpe * 0.65
    if n_obs < 252:  return annual_sharpe * 0.75
    return annual_sharpe * 0.85


def _haircut_sharpe(annual_sharpe: float, sample_days: float, exposure_pct: float) -> float:
    if annual_sharpe <= 0:
        return annual_sharpe
    sf = min(1.0, sample_days / 252.0)
    ef = min(1.0, max(exposure_pct / 100.0, 0.20))
    return annual_sharpe * math.sqrt(sf * ef)


# ==========================================================
# DATA LOADING
# ==========================================================

def load_set(month: str, paths: dict) -> dict | None:
    try:
        trades = pd.read_csv(paths["trades"])
        equity = pd.read_csv(paths["equity"])
        with open(paths["summary"], "r", encoding="utf-8") as f:
            summary = json.load(f)
        return {"month": month, "trades": trades, "equity": equity, "summary": summary}
    except FileNotFoundError as e:
        print(f"  [WARN] Missing file for {month}: {e}")
        return None


# ==========================================================
# COMPUTE METRICS PER SET
# ==========================================================

def compute_set_metrics(data: dict) -> dict:
    trades = data["trades"].copy()
    equity = data["equity"].copy()
    month  = data["month"]
    sm     = data["summary"].get("metrics", {})

    closed = trades[trades["status"] == "CLOSED"].copy() if "status" in trades.columns else trades.copy()
    pnl    = closed["pnl"].astype(float).values
    wins   = pnl[pnl > 0]
    losses = pnl[pnl < 0]

    eq = equity["equity"].astype(float).values
    peak = np.maximum.accumulate(eq)
    dd_pct = ((eq - peak) / peak)

    # Time
    ts = pd.to_datetime(equity["timestamp"], errors="coerce").dropna()
    backtest_days = (ts.iloc[-1] - ts.iloc[0]).total_seconds() / 86400.0

    # Daily returns
    daily_eq = (
        equity.assign(_ts=pd.to_datetime(equity["timestamp"], errors="coerce"))
        .dropna(subset=["_ts"])
        .set_index("_ts")["equity"]
        .resample("1D").last().dropna()
    )
    daily_ret = daily_eq.pct_change().dropna().values

    # Exposure
    exposure_pct = equity["position_open"].mean() * 100 if "position_open" in equity.columns else 0.0

    # XIRR (CAGR from single start/end point)
    net_profit = float(pnl.sum())
    ending_equity = float(eq[-1])
    total_return_pct = _safe_div(ending_equity - STARTING_CAPITAL, STARTING_CAPITAL) * 100
    years = backtest_days / 365.25
    xirr_pct = ((ending_equity / STARTING_CAPITAL) ** (1 / years) - 1) * 100 if years > 0 else 0.0

    # Sharpe
    sh_raw = _raw_sharpe(daily_ret)
    sh_ann = _annualized_sharpe(daily_ret, 252)
    sh_def = _deflated_sharpe(sh_ann, len(daily_ret))
    sh_hc  = _haircut_sharpe(sh_ann, backtest_days, exposure_pct)
    sh_dep = min(sh_def, sh_hc)

    # Trade Sharpe
    trade_ret_arr = closed["pnl_pct"].astype(float).values if "pnl_pct" in closed.columns else np.array([])
    t_raw = _raw_sharpe(trade_ret_arr)
    tpd   = _safe_div(len(closed), backtest_days)
    t_ann = _annualized_sharpe(trade_ret_arr, tpd * 252)
    t_def = _deflated_sharpe(t_ann, len(trade_ret_arr))

    # Annualized trade stats
    best_at = worst_at = median_at = top10_at = bot10_at = 0.0
    if len(closed) > 0 and "days_held" in closed.columns:
        hold = closed["days_held"].astype(float).clip(lower=1/48).values
        score = np.clip((trade_ret_arr / hold) * 100.0, -300, 300)
        best_at   = float(score.max())
        worst_at  = float(score.min())
        median_at = float(np.median(score))
        top10_at  = float(np.percentile(score, 90))
        bot10_at  = float(np.percentile(score, 10))

    # Monthly breakdown within the set
    closed2 = closed.copy()
    if "exit_time" in closed2.columns:
        closed2["exit_dt"] = pd.to_datetime(closed2["exit_time"], errors="coerce")
        closed2["ym"] = closed2["exit_dt"].dt.to_period("M").astype(str)
        monthly_pnl = closed2.groupby("ym")["pnl"].sum().to_dict()
        monthly_trades = closed2.groupby("ym")["pnl"].count().to_dict()
    else:
        monthly_pnl = {}
        monthly_trades = {}

    return {
        "month": month,
        "total_closed_trades": len(closed),
        "total_wins": len(wins),
        "total_losses": len(losses),
        "win_rate": _safe_div(len(wins), len(closed)),
        "gross_profit": float(wins.sum()),
        "gross_loss": float(losses.sum()),
        "net_profit": net_profit,
        "avg_win": float(wins.mean()) if len(wins) else 0.0,
        "avg_loss": float(abs(losses.mean())) if len(losses) else 0.0,
        "profit_factor": _safe_div(wins.sum(), abs(losses.sum())),
        "payoff_ratio": _safe_div(
            float(wins.mean()) if len(wins) else 0.0,
            float(abs(losses.mean())) if len(losses) else 0.0,
        ),
        "expectancy_per_trade": _safe_div(net_profit, len(closed)),
        "starting_capital": STARTING_CAPITAL,
        "ending_equity": ending_equity,
        "total_return_pct": total_return_pct,
        "xirr_pct": xirr_pct,
        "max_drawdown_pct": float(dd_pct.min() * 100),
        "max_drawdown_value": float(abs((eq - peak).min())),
        "backtest_days": backtest_days,
        "trades_per_day": tpd,
        "exposure_pct": exposure_pct,
        "active_day_ratio": float(np.mean(np.abs(daily_ret) > 1e-12) * 100) if len(daily_ret) else 0.0,
        "sharpe_raw": sh_raw,
        "sharpe_annualized": sh_ann,
        "sharpe_deflated": sh_def,
        "sharpe_haircut": sh_hc,
        "sharpe_deployable": sh_dep,
        "trade_sharpe_raw": t_raw,
        "trade_sharpe_annualized": t_ann,
        "trade_sharpe_deflated": t_def,
        "best_annualized_trade": best_at,
        "worst_annualized_trade": worst_at,
        "median_annualized_trade": median_at,
        "top10_annualized_trade": top10_at,
        "bottom10_annualized_trade": bot10_at,
        "monthly_pnl": monthly_pnl,
        "monthly_trades": monthly_trades,
        "score": data["summary"].get("score", 0),
        "grade": data["summary"].get("grade", "N/A"),
        "overfit_risk": data["summary"].get("overfit_report", {}).get("overfit_risk", "N/A"),
        "robustness_score": data["summary"].get("overfit_report", {}).get("robustness_score", 0),
        "trades_df": closed,
        "equity_df": equity,
        "daily_ret": daily_ret,
    }


# ==========================================================
# COMBINED METRICS (arithmetic — not compounded)
# ==========================================================

def compute_combined(all_metrics: list[dict]) -> dict:
    """
    Combine metrics across independent monthly backtests.
    Rules:
    - Counts, sums: simply additive
    - Ratios (win_rate, profit_factor, sharpe): computed from pooled data or simple mean
    - Return %: arithmetic average (NOT chained) — each month starts at 1500
    - XIRR: NOT combined (meaningless across independent ports)
    - Drawdown: worst individual month
    - Sharpe: pooled daily returns across all months
    """
    total_trades = sum(m["total_closed_trades"] for m in all_metrics)
    total_wins   = sum(m["total_wins"]  for m in all_metrics)
    total_losses = sum(m["total_losses"] for m in all_metrics)
    gross_profit = sum(m["gross_profit"] for m in all_metrics)
    gross_loss   = sum(m["gross_loss"]   for m in all_metrics)
    net_profit   = sum(m["net_profit"]   for m in all_metrics)
    total_days   = sum(m["backtest_days"] for m in all_metrics)

    # Pooled daily returns for Sharpe
    all_daily = np.concatenate([m["daily_ret"] for m in all_metrics if len(m["daily_ret"]) > 0])

    avg_exposure = np.mean([m["exposure_pct"] for m in all_metrics])
    sh_raw = _raw_sharpe(all_daily)
    sh_ann = _annualized_sharpe(all_daily, 252)
    sh_def = _deflated_sharpe(sh_ann, len(all_daily))
    sh_hc  = _haircut_sharpe(sh_ann, total_days, avg_exposure)
    sh_dep = min(sh_def, sh_hc)

    avg_return = np.mean([m["total_return_pct"] for m in all_metrics])
    best_month = max(m["total_return_pct"] for m in all_metrics)
    worst_month = min(m["total_return_pct"] for m in all_metrics)
    worst_dd = min(m["max_drawdown_pct"] for m in all_metrics)

    # Pooled trade-level Sharpe
    all_trade_ret = np.concatenate([
        m["trades_df"]["pnl_pct"].astype(float).values
        for m in all_metrics
        if "pnl_pct" in m["trades_df"].columns
    ])
    tpd_avg = _safe_div(total_trades, total_days)
    t_ann = _annualized_sharpe(all_trade_ret, tpd_avg * 252)
    t_def = _deflated_sharpe(t_ann, len(all_trade_ret))

    return {
        "total_closed_trades": total_trades,
        "total_wins": total_wins,
        "total_losses": total_losses,
        "win_rate": _safe_div(total_wins, total_trades),
        "gross_profit": gross_profit,
        "gross_loss": gross_loss,
        "net_profit": net_profit,
        "avg_win": _safe_div(gross_profit, total_wins),
        "avg_loss": _safe_div(abs(gross_loss), total_losses),
        "profit_factor": _safe_div(gross_profit, abs(gross_loss)),
        "expectancy_per_trade": _safe_div(net_profit, total_trades),
        "avg_monthly_return_pct": avg_return,
        "best_monthly_return_pct": best_month,
        "worst_monthly_return_pct": worst_month,
        "note_xirr": "Not applicable — each month is an independent backtest (port reset to 1500)",
        "worst_drawdown_pct": worst_dd,
        "total_days": total_days,
        "trades_per_day": tpd_avg,
        "avg_exposure_pct": avg_exposure,
        "sharpe_raw_pooled": sh_raw,
        "sharpe_annualized_pooled": sh_ann,
        "sharpe_deflated_pooled": sh_def,
        "sharpe_deployable_pooled": sh_dep,
        "trade_sharpe_deflated_pooled": t_def,
        "n_months": len(all_metrics),
        "profitable_months": sum(1 for m in all_metrics if m["net_profit"] > 0),
    }


# ==========================================================
# EXCEL BUILDER
# ==========================================================

def apply_header(ws, row, col, value, span=1, dark=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = FONT_HEADER if dark else FONT_SUBHDR
    c.fill = FILL_HEADER if dark else FILL_SUBHDR
    c.alignment = ALIGN_CENTER
    c.border = THIN_BORDER
    if span > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=col + span - 1
        )
    return c


def apply_cell(ws, row, col, value, fmt=None, bold=False, fill=None, align=None, color=None):
    c = ws.cell(row=row, column=col, value=value)
    font_args = {"name": "Arial", "size": 9, "bold": bold}
    if color:
        font_args["color"] = color
    c.font = Font(**font_args)
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    c.alignment = align or ALIGN_RIGHT
    c.border = THIN_BORDER
    return c


def pct_color(val):
    if val > 0:
        return "00B050"
    elif val < 0:
        return "FF4B4B"
    return "000000"


# ----------------------------------------------------------
# Sheet 1: Cover / Summary
# ----------------------------------------------------------

def build_cover(wb, all_metrics, combined):
    ws = wb.create_sheet("📊 Overview")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 22

    # Title
    ws.row_dimensions[1].height = 40
    ws.merge_cells("A1:F1")
    t = ws.cell(row=1, column=1, value="BACKTEST PERFORMANCE OVERVIEW — Jan to Apr 2026")
    t.font = Font(name="Arial", bold=True, size=14, color=C_WHITE)
    t.fill = FILL_HEADER
    t.alignment = ALIGN_CENTER
    t.border = THIN_BORDER

    ws.merge_cells("A2:F2")
    note = ws.cell(row=2, column=1,
        value="⚠️  Each month is an INDEPENDENT backtest. Portfolio resets to 1,500 THB each month. "
              "Results cannot be compounded or chained. Cross-month aggregates use arithmetic pooling only.")
    note.font = Font(name="Arial", italic=True, size=9, color="7F3F00")
    note.fill = PatternFill("solid", fgColor="FFF2CC")
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    note.border = THIN_BORDER
    ws.row_dimensions[2].height = 30

    # Column headers
    r = 4
    headers = ["Metric", "Jan-2026", "Feb-2026", "Mar-2026", "Apr-2026", "Combined (Pooled)"]
    for i, h in enumerate(headers, 1):
        apply_header(ws, r, i, h)
    ws.row_dimensions[r].height = 20

    MONTHS = [m["month"] for m in all_metrics]

    def add_section_header(row, label):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws.cell(row=row, column=1, value=label)
        c.font = Font(name="Arial", bold=True, size=9, color=C_WHITE)
        c.fill = PatternFill("solid", fgColor=C_ACCENT)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = THIN_BORDER
        ws.row_dimensions[row].height = 16

    def add_row(row, label, values_fn, fmt="#,##0.00", combined_val=None, combined_fmt=None):
        fill = FILL_ALT if row % 2 == 0 else FILL_WHITE
        apply_cell(ws, row, 1, label, bold=False, fill=fill, align=ALIGN_LEFT)
        for ci, m in enumerate(all_metrics, 2):
            v = values_fn(m)
            color = None
            if isinstance(v, float) and "%" in (fmt or ""):
                color = pct_color(v)
            apply_cell(ws, row, ci, v, fmt=fmt, fill=fill, color=color)
        # Combined column
        cv = combined_val if combined_val is not None else ""
        cfmt = combined_fmt or fmt
        apply_cell(ws, row, 6, cv, fmt=cfmt, fill=fill)
        ws.row_dimensions[row].height = 15

    r = 5
    add_section_header(r, "TRADE SUMMARY"); r += 1
    add_row(r, "Total Closed Trades",    lambda m: m["total_closed_trades"], "#,##0", combined["total_closed_trades"], "#,##0"); r += 1
    add_row(r, "Wins",                   lambda m: m["total_wins"],          "#,##0", combined["total_wins"], "#,##0"); r += 1
    add_row(r, "Losses",                 lambda m: m["total_losses"],        "#,##0", combined["total_losses"], "#,##0"); r += 1
    add_row(r, "Win Rate",               lambda m: m["win_rate"],            "0.00%", combined["win_rate"], "0.00%"); r += 1
    add_row(r, "Trades / Day",           lambda m: m["trades_per_day"],      "0.00",  combined["trades_per_day"], "0.00"); r += 1

    add_section_header(r, "PROFITABILITY"); r += 1
    add_row(r, "Gross Profit (THB)",     lambda m: m["gross_profit"],        "#,##0.00", combined["gross_profit"]); r += 1
    add_row(r, "Gross Loss (THB)",       lambda m: m["gross_loss"],          "#,##0.00", combined["gross_loss"]); r += 1
    add_row(r, "Net Profit (THB)",       lambda m: m["net_profit"],          "#,##0.00", combined["net_profit"]); r += 1
    add_row(r, "Avg Win (THB)",          lambda m: m["avg_win"],             "#,##0.00", combined["avg_win"]); r += 1
    add_row(r, "Avg Loss (THB)",         lambda m: m["avg_loss"],            "#,##0.00", combined["avg_loss"]); r += 1
    add_row(r, "Profit Factor",          lambda m: m["profit_factor"],       "0.000",   combined["profit_factor"]); r += 1
    add_row(r, "Payoff Ratio",           lambda m: m["payoff_ratio"],        "0.000"); r += 1
    add_row(r, "Expectancy / Trade (THB)", lambda m: m["expectancy_per_trade"], "#,##0.00", combined["expectancy_per_trade"]); r += 1

    add_section_header(r, "RETURN METRICS (per independent backtest)"); r += 1
    add_row(r, "Starting Capital (THB)", lambda m: m["starting_capital"],       "#,##0.00", "1,500.00 × 4 sets", "General"); r += 1
    add_row(r, "Ending Equity (THB)",    lambda m: m["ending_equity"],          "#,##0.00"); r += 1
    add_row(r, "Total Return % (monthly)", lambda m: m["total_return_pct"]/100, "0.00%",  combined["avg_monthly_return_pct"]/100, "0.00% avg"); r += 1
    add_row(r, "XIRR (annualized, per backtest)", lambda m: m["xirr_pct"]/100,  "0.00%",  "N/A — independent ports", "General"); r += 1
    add_row(r, "Best Monthly Return",    lambda m: "",                           "General", combined["best_monthly_return_pct"]/100, "0.00%"); r += 1
    add_row(r, "Worst Monthly Return",   lambda m: "",                           "General", combined["worst_monthly_return_pct"]/100, "0.00%"); r += 1

    add_section_header(r, "RISK"); r += 1
    add_row(r, "Max Drawdown %",           lambda m: m["max_drawdown_pct"]/100,  "0.00%", combined["worst_drawdown_pct"]/100, "0.00% (worst)"); r += 1
    add_row(r, "Max Drawdown Value (THB)", lambda m: m["max_drawdown_value"],     "#,##0.00"); r += 1
    add_row(r, "Exposure Time %",          lambda m: m["exposure_pct"]/100,      "0.00%", combined["avg_exposure_pct"]/100, "0.00% avg"); r += 1

    add_section_header(r, "SHARPE FRAMEWORK"); r += 1
    add_row(r, "Sharpe Raw (daily)",     lambda m: m["sharpe_raw"],          "0.000", combined["sharpe_raw_pooled"]); r += 1
    add_row(r, "Sharpe Annualized",      lambda m: m["sharpe_annualized"],   "0.000", combined["sharpe_annualized_pooled"]); r += 1
    add_row(r, "Sharpe Deflated",        lambda m: m["sharpe_deflated"],     "0.000", combined["sharpe_deflated_pooled"]); r += 1
    add_row(r, "Sharpe Haircut",         lambda m: m["sharpe_haircut"],      "0.000"); r += 1
    add_row(r, "Sharpe Deployable",      lambda m: m["sharpe_deployable"],   "0.000", combined["sharpe_deployable_pooled"]); r += 1
    add_row(r, "Trade Sharpe (deflated)", lambda m: m["trade_sharpe_deflated"], "0.000", combined["trade_sharpe_deflated_pooled"]); r += 1

    add_section_header(r, "ANNUALIZED TRADE EFFICIENCY"); r += 1
    add_row(r, "Best Annualized Trade %",     lambda m: m["best_annualized_trade"]/100,    "0.00%"); r += 1
    add_row(r, "Worst Annualized Trade %",    lambda m: m["worst_annualized_trade"]/100,   "0.00%"); r += 1
    add_row(r, "Median Annualized Trade %",   lambda m: m["median_annualized_trade"]/100,  "0.00%"); r += 1
    add_row(r, "Top 10% Annualized Trade",    lambda m: m["top10_annualized_trade"]/100,   "0.00%"); r += 1
    add_row(r, "Bottom 10% Annualized Trade", lambda m: m["bottom10_annualized_trade"]/100,"0.00%"); r += 1

    add_section_header(r, "MODEL & OVERFIT"); r += 1
    add_row(r, "Run Score",             lambda m: m["score"],               "0.00"); r += 1
    add_row(r, "Grade",                 lambda m: m["grade"],               "General"); r += 1
    add_row(r, "Overfit Risk",          lambda m: m["overfit_risk"],        "General"); r += 1
    add_row(r, "Robustness Score",      lambda m: m["robustness_score"],    "0.00"); r += 1
    add_row(r, "Backtest Days",         lambda m: m["backtest_days"],       "0.0", combined["total_days"]); r += 1


# ----------------------------------------------------------
# Sheet 2: Monthly Returns Series
# ----------------------------------------------------------

def build_monthly_returns(wb, all_metrics):
    ws = wb.create_sheet("📅 Monthly Returns")
    ws.sheet_view.showGridLines = False

    for col, w in zip("ABCDE", [18, 14, 14, 14, 14]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:E1")
    t = ws.cell(row=1, column=1, value="MONTHLY RETURN SERIES (Each month = independent backtest from 1,500 THB)")
    t.font = Font(name="Arial", bold=True, size=12, color=C_WHITE)
    t.fill = FILL_HEADER
    t.alignment = ALIGN_CENTER
    t.border = THIN_BORDER
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:E2")
    note = ws.cell(row=2, column=1,
        value="NOTE: Return % values are MONTHLY (not annual, not compounded). "
              "Do not chain or multiply across months.")
    note.font = FONT_WARN
    note.fill = FILL_WARN
    note.alignment = Alignment(horizontal="left", wrap_text=True)
    note.border = THIN_BORDER
    ws.row_dimensions[2].height = 24

    headers = ["Month", "Net Profit (THB)", "Return % (of 1500)", "Trades", "Win Rate"]
    for i, h in enumerate(headers, 1):
        apply_header(ws, 3, i, h)

    r = 4
    for m in all_metrics:
        fill = FILL_ALT if r % 2 == 0 else FILL_WHITE
        apply_cell(ws, r, 1, m["month"], fill=fill, align=ALIGN_LEFT)
        apply_cell(ws, r, 2, m["net_profit"], fmt="#,##0.00", fill=fill, color=pct_color(m["net_profit"]))
        apply_cell(ws, r, 3, m["total_return_pct"] / 100, fmt="0.00%", fill=fill, color=pct_color(m["total_return_pct"]))
        apply_cell(ws, r, 4, m["total_closed_trades"], fmt="#,##0", fill=fill)
        apply_cell(ws, r, 5, m["win_rate"], fmt="0.00%", fill=fill)
        ws.row_dimensions[r].height = 15
        r += 1

    # Totals row
    fill = PatternFill("solid", fgColor="D9E1F2")
    apply_cell(ws, r, 1, "TOTAL / AVG", bold=True, fill=fill, align=ALIGN_LEFT)
    apply_cell(ws, r, 2, sum(m["net_profit"] for m in all_metrics), fmt="#,##0.00", bold=True, fill=fill)
    avg_ret = np.mean([m["total_return_pct"] for m in all_metrics])
    c = ws.cell(row=r, column=3, value=avg_ret / 100)
    c.font = Font(name="Arial", bold=True, size=9)
    c.number_format = "0.00%"
    c.fill = fill
    c.alignment = ALIGN_RIGHT
    c.border = THIN_BORDER
    # Note cell
    nc = ws.cell(row=r+1, column=1, value="← avg monthly return (arithmetic mean, NOT compounded)")
    nc.font = FONT_WARN
    ws.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=5)


# ----------------------------------------------------------
# Sheet 3: All Trades (raw)
# ----------------------------------------------------------

def build_all_trades(wb, all_metrics):
    ws = wb.create_sheet("📋 All Trades")
    ws.sheet_view.showGridLines = False

    keep_cols = [
        "entry_time", "exit_time", "status", "side", "qty",
        "entry_price", "exit_price", "entry_notional", "exit_value",
        "pnl", "pnl_pct", "days_held", "minutes_held",
        "reason", "session_name",
        "entry_prob_buy", "entry_prob_sell", "mfe", "mae"
    ]

    dfs = []
    for m in all_metrics:
        df = m["trades_df"].copy()
        df.insert(0, "backtest_month", m["month"])
        dfs.append(df)

    combined_df = pd.concat(dfs, ignore_index=True)
    cols_present = ["backtest_month"] + [c for c in keep_cols if c in combined_df.columns]
    combined_df = combined_df[cols_present]

    col_widths = {
        "backtest_month": 14, "entry_time": 18, "exit_time": 18,
        "status": 10, "side": 8, "qty": 10, "entry_price": 12,
        "exit_price": 12, "entry_notional": 14, "exit_value": 12,
        "pnl": 10, "pnl_pct": 10, "days_held": 10, "minutes_held": 10,
        "reason": 16, "session_name": 18,
        "entry_prob_buy": 14, "entry_prob_sell": 14, "mfe": 10, "mae": 10
    }

    for i, col in enumerate(cols_present, 1):
        ws.column_dimensions[get_column_letter(i)].width = col_widths.get(col, 12)
        apply_header(ws, 1, i, col.replace("_", " ").title())

    pnl_col_idx = cols_present.index("pnl") + 1 if "pnl" in cols_present else None

    for ri, (_, row) in enumerate(combined_df.iterrows(), 2):
        fill = FILL_ALT if ri % 2 == 0 else FILL_WHITE
        for ci, col in enumerate(cols_present, 1):
            val = row[col]
            if pd.isna(val):
                val = ""
            fmt = None
            color = None
            if col in ("pnl", "entry_notional", "exit_value"):
                fmt = "#,##0.00"
                if col == "pnl":
                    color = pct_color(float(val)) if val != "" else None
            elif col == "pnl_pct":
                fmt = "0.0000%"
            elif col in ("entry_price", "exit_price"):
                fmt = "#,##0.00"
            elif col in ("days_held", "minutes_held"):
                fmt = "0.00"
            elif col in ("entry_prob_buy", "entry_prob_sell"):
                fmt = "0.000"
            elif col in ("mfe", "mae"):
                fmt = "#,##0.00"
            apply_cell(ws, ri, ci, val, fmt=fmt, fill=fill, color=color,
                       align=ALIGN_LEFT if col in ("entry_time","exit_time","reason","session_name","backtest_month","side","status") else ALIGN_RIGHT)
        ws.row_dimensions[ri].height = 14

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols_present))}1"


# ----------------------------------------------------------
# Sheet 4: Equity Curves
# ----------------------------------------------------------

def build_equity_curves(wb, all_metrics):
    ws = wb.create_sheet("📈 Equity Curves")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:B1")
    t = ws.cell(row=1, column=1, value="EQUITY CURVES — Each month independent (resets to 1,500 THB)")
    t.font = Font(name="Arial", bold=True, size=12, color=C_WHITE)
    t.fill = FILL_HEADER
    t.alignment = ALIGN_CENTER
    t.border = THIN_BORDER
    ws.row_dimensions[1].height = 30

    start_col = 1
    for m in all_metrics:
        eq_df = m["equity_df"][["timestamp", "equity"]].copy()
        eq_df.columns = [f"timestamp_{m['month']}", f"equity_{m['month']}"]
        eq_df[f"drawdown_pct_{m['month']}"] = (
            (eq_df[f"equity_{m['month']}"] - eq_df[f"equity_{m['month']}"].cummax())
            / eq_df[f"equity_{m['month']}"].cummax() * 100
        )
        ncols = len(eq_df.columns)

        for ci, col in enumerate(eq_df.columns, start_col):
            apply_header(ws, 2, ci, col.replace("_", " ").title(), dark=False)
            ws.column_dimensions[get_column_letter(ci)].width = 18

        for ri, (_, row) in enumerate(eq_df.iterrows(), 3):
            fill = FILL_ALT if ri % 2 == 0 else FILL_WHITE
            for ci, col in enumerate(eq_df.columns, start_col):
                val = row[col]
                fmt = None
                if "equity" in col:
                    fmt = "#,##0.00"
                elif "drawdown" in col:
                    fmt = "0.00%"
                    val = val / 100 if pd.notna(val) else val
                apply_cell(ws, ri, ci, val if pd.notna(val) else "", fmt=fmt, fill=fill)
            ws.row_dimensions[ri].height = 12

        start_col += ncols + 1  # gap between months


# ----------------------------------------------------------
# Sheet 5: Summary Stats CSV-friendly
# ----------------------------------------------------------

def build_summary_csv_sheet(wb, all_metrics, combined):
    ws = wb.create_sheet("📥 Export-Ready Summary")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:C1")
    t = ws.cell(row=1, column=1, value="EXPORT-READY FLAT SUMMARY (suitable for slides / further analysis)")
    t.font = Font(name="Arial", bold=True, size=11, color=C_WHITE)
    t.fill = FILL_HEADER
    t.alignment = ALIGN_CENTER
    t.border = THIN_BORDER
    ws.row_dimensions[1].height = 25

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16

    headers = ["Metric", "Value", "Unit / Note"]
    for i, h in enumerate(headers, 1):
        apply_header(ws, 2, i, h)

    rows = []
    # Per month
    for m in all_metrics:
        rows.append((f"[{m['month']}] Net Profit", round(m["net_profit"], 4), "THB"))
        rows.append((f"[{m['month']}] Return %", round(m["total_return_pct"], 4), "% of 1,500 THB capital"))
        rows.append((f"[{m['month']}] XIRR", round(m["xirr_pct"], 4), "% annualized (independent)"))
        rows.append((f"[{m['month']}] Win Rate", round(m["win_rate"] * 100, 2), "%"))
        rows.append((f"[{m['month']}] Profit Factor", round(m["profit_factor"], 4), "x"))
        rows.append((f"[{m['month']}] Sharpe Deployable", round(m["sharpe_deployable"], 4), ""))
        rows.append((f"[{m['month']}] Max Drawdown", round(m["max_drawdown_pct"], 4), "%"))
        rows.append((f"[{m['month']}] Trades", m["total_closed_trades"], "count"))
        rows.append((f"[{m['month']}] Score / Grade", f"{m['score']} / {m['grade']}", ""))
        rows.append(("---", "---", "---"))

    # Combined
    rows.append(("COMBINED: Total Trades", combined["total_closed_trades"], "count (pooled)"))
    rows.append(("COMBINED: Total Net Profit", round(combined["net_profit"], 4), "THB (sum)"))
    rows.append(("COMBINED: Avg Monthly Return", round(combined["avg_monthly_return_pct"], 4), "% (arithmetic avg)"))
    rows.append(("COMBINED: Best Month Return", round(combined["best_monthly_return_pct"], 4), "%"))
    rows.append(("COMBINED: Worst Month Return", round(combined["worst_monthly_return_pct"], 4), "%"))
    rows.append(("COMBINED: Win Rate", round(combined["win_rate"] * 100, 2), "% (pooled)"))
    rows.append(("COMBINED: Profit Factor", round(combined["profit_factor"], 4), "x (pooled)"))
    rows.append(("COMBINED: Sharpe Deployable (pooled)", round(combined["sharpe_deployable_pooled"], 4), "pooled daily returns"))
    rows.append(("COMBINED: Worst Drawdown", round(combined["worst_drawdown_pct"], 4), "% (worst individual month)"))
    rows.append(("COMBINED: Profitable Months", combined["profitable_months"], f"out of {combined['n_months']}"))
    rows.append(("COMBINED: XIRR", "N/A", "Cannot combine — independent ports"))

    for ri, (label, val, unit) in enumerate(rows, 3):
        fill = FILL_ALT if ri % 2 == 0 else FILL_WHITE
        apply_cell(ws, ri, 1, label, fill=fill, align=ALIGN_LEFT)
        apply_cell(ws, ri, 2, val, fill=fill)
        apply_cell(ws, ri, 3, unit, fill=fill, align=ALIGN_LEFT)
        ws.row_dimensions[ri].height = 14


# ==========================================================
# MAIN
# ==========================================================

def main():
    print("Loading backtest data...")
    raw_sets = []
    for month, paths in SETS.items():
        data = load_set(month, paths)
        if data:
            raw_sets.append(data)
            print(f"  ✓ {month} loaded")
        else:
            print(f"  ✗ {month} skipped (files not found)")

    if not raw_sets:
        print("ERROR: No data loaded. Check paths.")
        return

    print("Computing metrics...")
    all_metrics = [compute_set_metrics(d) for d in raw_sets]
    combined = compute_combined(all_metrics)

    print("Building Excel report...")
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_cover(wb, all_metrics, combined)
    build_monthly_returns(wb, all_metrics)
    build_all_trades(wb, all_metrics)
    build_equity_curves(wb, all_metrics)
    build_summary_csv_sheet(wb, all_metrics, combined)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"\n✅  Saved: {OUTPUT_PATH}")

    # Also save a flat CSV for quick use
    csv_rows = []
    for m in all_metrics:
        csv_rows.append({
            "month": m["month"],
            "total_trades": m["total_closed_trades"],
            "wins": m["total_wins"],
            "losses": m["total_losses"],
            "win_rate_pct": round(m["win_rate"] * 100, 2),
            "net_profit_thb": round(m["net_profit"], 4),
            "total_return_pct": round(m["total_return_pct"], 4),
            "xirr_pct": round(m["xirr_pct"], 4),
            "gross_profit": round(m["gross_profit"], 4),
            "gross_loss": round(m["gross_loss"], 4),
            "avg_win_thb": round(m["avg_win"], 4),
            "avg_loss_thb": round(m["avg_loss"], 4),
            "profit_factor": round(m["profit_factor"], 4),
            "payoff_ratio": round(m["payoff_ratio"], 4),
            "expectancy_per_trade": round(m["expectancy_per_trade"], 4),
            "ending_equity": round(m["ending_equity"], 4),
            "max_drawdown_pct": round(m["max_drawdown_pct"], 4),
            "max_drawdown_value": round(m["max_drawdown_value"], 4),
            "backtest_days": round(m["backtest_days"], 2),
            "trades_per_day": round(m["trades_per_day"], 4),
            "exposure_pct": round(m["exposure_pct"], 4),
            "active_day_ratio": round(m["active_day_ratio"], 4),
            "sharpe_raw": round(m["sharpe_raw"], 4),
            "sharpe_annualized": round(m["sharpe_annualized"], 4),
            "sharpe_deflated": round(m["sharpe_deflated"], 4),
            "sharpe_haircut": round(m["sharpe_haircut"], 4),
            "sharpe_deployable": round(m["sharpe_deployable"], 4),
            "trade_sharpe_raw": round(m["trade_sharpe_raw"], 4),
            "trade_sharpe_annualized": round(m["trade_sharpe_annualized"], 4),
            "trade_sharpe_deflated": round(m["trade_sharpe_deflated"], 4),
            "best_annualized_trade_pct": round(m["best_annualized_trade"], 4),
            "worst_annualized_trade_pct": round(m["worst_annualized_trade"], 4),
            "median_annualized_trade_pct": round(m["median_annualized_trade"], 4),
            "top10_annualized_trade_pct": round(m["top10_annualized_trade"], 4),
            "bottom10_annualized_trade_pct": round(m["bottom10_annualized_trade"], 4),
            "score": m["score"],
            "grade": m["grade"],
            "overfit_risk": m["overfit_risk"],
            "robustness_score": m["robustness_score"],
            "note": "independent_backtest_port_reset_1500thb_per_month",
        })

    csv_path = BASE / "backtest_monthly_summary_2026.csv"
    pd.DataFrame(csv_rows).to_csv(csv_path, index=False)
    print(f"✅  Saved CSV: {csv_path}")


if __name__ == "__main__":
    main()